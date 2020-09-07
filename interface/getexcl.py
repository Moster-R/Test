#!/usr/bin/env python
# _*_ coding: utf-8 _*_
# @Time : 2020/9/4 22:45
# @Author : 墨
# @Version：V 0.1
# @File : getexcl.py
# @desc :

import openpyxl
import requests
import os
class ExclMth():

    def __init__(self,file_name,sheet_name):

        self.file_name = file_name
        self.wookbook = openpyxl.load_workbook(self.file_name)

        # 获取工作表对象
        self.sheet = self.wookbook[sheet_name]

        # 获取最大行
        self.max_Row = self.sheet.max_row
        # 获取列
        self.max_column = self.sheet.max_column
        # 获取总得行数
        # self.row = self.sheet.max_row

    def ReadExcl(self):
        dataList=[]
        try:
            for row in self.sheet.rows:
                tempList =[]
                for cell in row:   # 取出第一行就是row[1:]
                    tempList.append(cell.value)
                dataList.append(tempList)
        except:
            print("%s加载失败"% self.file_name)

        else:
            return dataList[1:]

    def SaveExcl(self,row,text):
        try:
            self.sheet.cell(row,self.max_column,text)
            self.wookbook.save(self.file_name)
        except:
            print("%s保存失败"% self.file_name)




    def visa_json(method,url,params = None,data = None,json = None,**kwargs):

        response = requests.request(
                method,
                url,
                params = params,
                data = data,
                json = json,
                **kwargs
        )
        try:
            return response.json()
        except:
            print("不是json数据")
            return None



# if __name__ == '__main__':
#
#     dir_url = os.path.join(os.getcwd()+r'..\..\excl\test_case_api.xlsx')
#     excl = ExclMth(dir_url,'login')
#     data = excl.ReadExcl()
#     excl.SaveExcl(7,"pass")
